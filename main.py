from et_xmlfile import xmlfile
import jdcal
from kivy.core.window import Window

from kivy.uix.image import Image, AsyncImage
from kivy.graphics import Rectangle, Color
from kivy.uix.spinner import Spinner
from openpyxl import Workbook
import openpyxl
import os
from kivy.app import App
from kivy.uix.button import Button
from kivy.graphics.instructions import Canvas
#python3,kivy==master,openpyxl,et_xmlfile,jsonrpclib,jsonpickle,jdcal
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.widget import Widget
class MainApp(App):
    def build(self):

        print('os getcwd'+os.getcwd())
        try:
            wb = openpyxl.load_workbook("datafile.xlsx")

        except:
            wb = openpyxl.Workbook()
            w = wb.create_sheet("Sheet1")

            wb.save(filename='datafile.xlsx')
        self.main_layout = FloatLayout(size=(50, 50))


        button1 = Button(
            text='Add Item'
            ,background_color=(0,255,0,0.7)

            ,


            pos=(20, 700),
            size_hint=(.50, .1),
            pos_hint={'x': .25, 'y': .85}
        )
        try:
            Window.clearcolor = (1, 1, 1, 1)
            a = Image(source='ar1.png',size=(1200,900))
            self.main_layout.add_widget(a)
        except:
            pass

        button1.bind(on_press=self.on_button_press1)
        self.main_layout.add_widget(button1)
        button2 = Button(
            text='Search Item',
            background_color = (0, 255, 0, 0.7),
            pos=(20, 700),
            size_hint=(.50, .1),
            pos_hint={'x': .25, 'y': .65}
        )
        button2.bind(on_press=self.on_button_press2)
        self.main_layout.add_widget(button2)
        button3 = Button(
            text='All Items',
            pos=(20, 700),background_color = (0, 255, 0, 0.7),
            size_hint=(.50, .1),
            pos_hint={'x': .25, 'y': .45}
        )
        button3.bind(on_press=self.on_button_press3)
        self.main_layout.add_widget(button3)

        button4 = Button(
            text='Total Items',
            pos=(20, 700),background_color = (0, 255, 0, 0.7),
            size_hint=(.50, .1),
            pos_hint={'x': .25, 'y': .25}
        )
        button5 = Button(
            text='Delete Item',
            pos=(20, 700),background_color = (0, 255, 0, 0.7),
            size_hint=(.50, .1),
            pos_hint={'x': .25, 'y': .05}
        )
        button4.bind(on_press=self.on_button_press4)
        self.main_layout.add_widget(button4)
        button5.bind(on_press=self.on_button_press5)
        self.main_layout.add_widget(button5)
        return self.main_layout

    def on_button_press1(self, instance):
        self.main_layout.clear_widgets()
        app1 = additem1()
        app1.run()
        print('hi')

    def on_button_press2(self, instance):
        self.main_layout.clear_widgets()
        f = search_item_class()
        f.run()

    def on_button_press3(self, instance):
        self.main_layout.clear_widgets()
        f = spin()
        f.run()

    def on_button_press4(self, instance):
        self.main_layout.clear_widgets()
        f = total_item_class()
        f.run()
    def on_button_press5(self, instance):
        self.main_layout.clear_widgets()
        f = delete_item_class()
        f.run()
class delete_record():
    def __init__(self,given_str):

        given_str=str(given_str)
        given_str=given_str.strip()

        wb = openpyxl.load_workbook("datafile.xlsx")
        sheet = wb['Sheet1']  # wb.get_sheet_names())

        columnvalue=1
        for i in range(1, sheet.max_row + 1):
            if ((sheet.cell(row=i, column=columnvalue).value == given_str)):

                cell = sheet.cell(row=i, column=1)
                cell.value = None
                cell = sheet.cell(row=i, column=2)
                cell.value =  None

                break
            else:
                pass
        wb.save("datafile.xlsx")


class delete_item_class1(App):
    def build(self):

        self.main_layout2 = FloatLayout(size=(50, 50))
        try:
            Window.clearcolor = (1, 1, 1, 1)
            a = Image(source='ar1.png',size=(1200,900))
            self.main_layout2.add_widget(a)
        except:
            pass
        self.item = delete_item_class().data
        obj = words_filter(self.item)
        self.item = obj.output()
        obj2 = delete_record(self.item)


        string = 'Done'
        label1 = Label(text=string,font_size=60,
                      size_hint=(.20, .10),color=[255,255, 255, 1],
                      pos_hint={'x': .40, 'y': .75})
        self.main_layout2.add_widget(label1)

        button1 = Button(
            text='Ok',background_color = (0, 255, 0, 0.7),

        size_hint=(.20,.08),
            pos_hint={'x': .45, 'y': .10}
        )
        button1.bind(on_press=self.ok)
        self.main_layout2.add_widget(button1)


        return self.main_layout2
    def ok(self,instance):

        self.main_layout2.clear_widgets()
        f=delete_record(self.item)
        f=delete_item_class()
        f.run()

class delete_item_class(App):
    data=''
    def build(self):

        self.main_layout2 = FloatLayout(size=(50, 50))
        try:
            Window.clearcolor = (1, 1, 1, 1)
            a = Image(source='ar1.png',size=(1200,900))
            self.main_layout2.add_widget(a)
        except:
            pass
        label1 = Label(text='item name',color=[255,255, 255, 1],
                      size_hint=(.20, .10),
                      pos_hint={'x': .2, 'y': .70})
        self.main_layout2.add_widget(label1)
        self.textbox1 = TextInput(

            multiline=False, readonly=False, font_size=40, size_hint = (.35, .1), pos_hint={'x': .50, 'y': .70}

        )
        self.main_layout2.add_widget(self.textbox1)


        button1 = Button(
            text='delete',background_color = (0, 255, 0, 0.7),

        size_hint=(.20,.08),
            pos_hint={'x': .30, 'y': .30}
        )
        button1.bind(on_press=self.delete)
        self.main_layout2.add_widget(button1)
        button2 = Button(
            text='Back',background_color = (0, 255, 0, 0.7),

            size_hint=(.20, .08),
            pos_hint={'x': .55, 'y': .30}
        )
        button2.bind(on_press=self.back)

        #self.objret = textbox(self.textbox1.text,self.textbox2.text)
        self.main_layout2.add_widget(button2)

        return self.main_layout2
    def back(self,instance):

        self.main_layout2.clear_widgets()
        f=MainApp()
        f.run()


    def delete(self, instance):
        delete_item_class.data = self.textbox1.text
        self.main_layout2.clear_widgets()
        f=delete_item_class1()
        f.run()
class total_item_class(App):
    def build(self):
        self.main_layout2 = FloatLayout(size=(50, 50))
        try:
            Window.clearcolor = (1, 1, 1, 1)
            a = Image(source='ar1.png',size=(1200,900))
            self.main_layout2.add_widget(a)
        except:
            pass
        obj = itemslist()

        self.item = obj.get()

        string = 'Total number of records are : %s' % (len(self.item))
        label1 = Label(text=string, font_size=35,color=[255,255, 255, 1],
                       size_hint=(.20, .10),
                       pos_hint={'x': .40, 'y': .70})
        self.main_layout2.add_widget(label1)
        button1 = Button(
            text='Ok',background_color = (0, 255, 0, 0.7),

            size_hint=(.20,.08),
            pos_hint={'x': .45, 'y': .30}
        )
        button1.bind(on_press=self.ok)
        self.main_layout2.add_widget(button1)

        return self.main_layout2

    def ok(self, instance):
        self.main_layout2.clear_widgets()
        f = MainApp()
        f.run()
class search_item_class1(App):
    def build(self):

        self.main_layout2 = FloatLayout(size=(50, 50))
        try:
            Window.clearcolor = (1, 1, 1, 1)
            a = Image(source='ar1.png',size=(1200,900))
            self.main_layout2.add_widget(a)
        except:
            pass
        self.item = search_item_class().data
        obj = words_filter(self.item)
        self.item = obj.output()
        obj2 = record_availability_checker(self.item)
        self.value = str(obj2.get())


        string = '%s is at %s' % (self.item,self.value)
        label1 = Label(text=string,font_size=40,color=[255,255, 255, 1],
                      size_hint=(.20, .10),
                      pos_hint={'x': .40, 'y': .75})
        self.main_layout2.add_widget(label1)

        button1 = Button(
            text='Ok',background_color = (0, 255, 0, 0.7),

        size_hint=(.20,.08),
            pos_hint={'x': .45, 'y': .10}
        )
        button1.bind(on_press=self.ok)
        self.main_layout2.add_widget(button1)


        return self.main_layout2
    def ok(self,instance):

        self.main_layout2.clear_widgets()
        f=search_item_class()
        f.run()

class search_item_class(App):
    data = ''


    def build(self):

        self.main_layout2 = FloatLayout(size=(50, 50))
        try:
            Window.clearcolor = (1, 1, 1, 1)
            a = Image(source='ar1.png',size=(1200,900))
            self.main_layout2.add_widget(a)
        except:
            pass
        label1 = Label(text='item name',color=[255,255, 255, 1],
                      size_hint=(.20, .10),
                      pos_hint={'x': .2, 'y': .70})
        self.main_layout2.add_widget(label1)
        self.textbox1 = TextInput(

            multiline=False, readonly=False, font_size=40, size_hint = (.35, .1), pos_hint={'x': .50, 'y': .70}

        )
        self.main_layout2.add_widget(self.textbox1)


        button1 = Button(
            text='search',background_color = (0, 255, 0, 0.7),

        size_hint=(.20,.08),
            pos_hint={'x': .30, 'y': .30}
        )
        button1.bind(on_press=self.search)
        self.main_layout2.add_widget(button1)
        button2 = Button(
            text='Back',background_color = (0, 255, 0, 0.7),

            size_hint=(.20,.08),
            pos_hint={'x': .55, 'y': .30}
        )
        button2.bind(on_press=self.back)

        #self.objret = textbox(self.textbox1.text,self.textbox2.text)
        self.main_layout2.add_widget(button2)

        return self.main_layout2
    def back(self,instance):

        self.main_layout2.clear_widgets()
        f=MainApp()
        f.run()


    def search(self, instance):
        search_item_class.data = self.textbox1.text
        self.main_layout2.clear_widgets()
        f=search_item_class1()
        f.run()


class spin(App):
    textbox1 = ""
    textbox2 = ""

    def build(self):
        obj = itemslist()
        self.items=obj.get()
        try:
            self.items.remove(None)
        except:
            pass
        print(obj.get())
        self.main_layout2 = FloatLayout(size=(50, 50))
        try:
            Window.clearcolor = (1, 1, 1, 1)
            a = Image(source='ar1.png',size=(1200,900))
            self.main_layout2.add_widget(a)
        except:
            pass
        self.spinner = Spinner(
            text='Click here',background_color = (0, 255, 0, 0.7),
            values=self.items,
            size_hint=(.20, .08),
            pos_hint={'x': .25, 'y': .80})

        self.main_layout2.add_widget(self.spinner)


        button1 = Button(
            text='Back',background_color = (0, 255, 0, 0.7),

            size_hint=(.20,.08),
            pos_hint={'x': .55, 'y': .80}
        )
        button1.bind(on_press=self.back)
        self.main_layout2.add_widget(button1)


        return self.main_layout2

    def back(self, instance):

        self.main_layout2.clear_widgets()
        #time.sleep(.5)
        #pyautogui.click(50,50)
        f = MainApp()
        f.run()





class donepopupclass(App):

    def build(self):

        self.main_layout2 = FloatLayout(size=(50, 50))
        try:
            Window.clearcolor = (1, 1, 1, 1)
            a = Image(source='ar1.png',size=(1200,900))
            self.main_layout2.add_widget(a)
        except:
            pass
        label1 = Label(text='Done',font_size=60,color=[255,255, 255, 1],
                       size_hint=(.20, .10),
                       pos_hint={'x': .40, 'y': .70})
        self.main_layout2.add_widget(label1)



        button1 = Button(
            text='Ok',background_color = (0, 255, 0, 0.7),

            size_hint=(.20, .08),
            pos_hint={'x': .35, 'y': .30}
        )
        button1.bind(on_press=self.ok)

        self.main_layout2.add_widget(button1)

        return self.main_layout2

    def ok(self, instance):

        self.main_layout2.clear_widgets()
        f = additem1()
        f.run()




class additem1(App):

    def build(self):
        self.main_layout2 = FloatLayout(size=(50, 50))
        try:
            Window.clearcolor = (1, 1, 1, 1)
            a = Image(source='ar1.png',size=(1200,900))
            self.main_layout2.add_widget(a)
        except:
            pass
        label1 = Label(text='Enter name of the item',
                       size_hint=(.50, .1),color=[255,255, 255, 1],
                       pos_hint={'x': .25, 'y': .80})
        self.main_layout2.add_widget(label1)
        self.textbox1 = TextInput(

            multiline=False, readonly=False, size_hint=(.50,.05), pos_hint={'x': .25, 'y': .77}

        )
        self.main_layout2.add_widget(self.textbox1)
        label2 = Label(text='Enter name of the location',
                       size_hint=(.50, .1),color=[255,255, 255, 1],
                       pos_hint={'x': .25, 'y': .65})
        self.main_layout2.add_widget(label2)
        self.textbox2 = TextInput(

            multiline=False, readonly=False, size_hint=(.50, .05), pos_hint={'x': .25, 'y': .62}

        )
        self.main_layout2.add_widget(self.textbox2)


        button2 = Button(
            text='Submit',background_color = (0, 255, 0, 0.7),

            size_hint=(.20,.08),
            pos_hint={'x': .40, 'y': .30}
        )
        button2.bind(on_press=self.submit)
        button3 = Button(
            text='Back',background_color = (0, 255, 0, 0.7),

            size_hint=(.20, .08),
            pos_hint={'x': .40, 'y': .17}
        )
        button3.bind(on_press=self.back)
        self.main_layout2.add_widget(button3)
        # self.objret = textbox(self.textbox1.text,self.textbox2.text)
        self.main_layout2.add_widget(button2)

        return self.main_layout2

    def back(self, instance):
        self.main_layout2.clear_widgets()
        f = MainApp()
        f.run()

    def submit(self, instance):
        ob = words_filter(self.textbox1.text)
        k = ob.output()
        obj = record_availability_checker(k)
        temp1 = obj.get()
        x = self.textbox1.text
        y = self.textbox2.text
        ob1 = words_filter(x)

        x = ob1.output()
        ob2 = words_filter(y)
        y = ob2.output()
        additem1.textbox1 = x
        additem1.textbox2 = y

        if temp1 == "":

            if x != "" and y != "":
                self.main_layout2.clear_widgets()

                obj2 = record_add(x, y)
                f = donepopupclass()
                f.run()
            else:
                self.main_layout2.clear_widgets()
                f = additem1()
                f.run()



        else:
            self.main_layout2.clear_widgets()
            f = mypopup()
            f.run()
class mypopup(App):
    def build(self):
        self.main_layout = FloatLayout(size=(50, 50))
        try:
            Window.clearcolor = (1, 1, 1, 1)
            a = Image(source='ar1.png',size=(1200,900))
            self.main_layout.add_widget(a)
        except:
            pass
        ob = record_availability_checker(additem1.textbox1.strip())
        label = Label(text='           %s is already present at %s' %(additem1.textbox1.strip(),ob.get()),color=[255,255, 255, 1],size_hint=(.15, .1),
            pos_hint={'x': .2, 'y': .70})
        self.main_layout.add_widget(label)
        button1 = Button(
            text='continue',background_color = (0, 255, 0, 0.7),
            pos=(20, 700),
            size_hint=(.20,.08),
            pos_hint={'x': .2, 'y': .30}
        )
        button1.bind(on_press=self.on_button_press1)
        self.main_layout.add_widget(button1)
        button2 = Button(
            text='cancel',background_color = (0, 255, 0, 0.7),
            pos=(20, 700),
            size_hint=(.20,.08),
            pos_hint={'x': .50, 'y': .30}
        )
        button2.bind(on_press=self.on_button_press2)
        self.main_layout.add_widget(button2)
        return self.main_layout
    def on_button_press1(self, instance):

        x=additem1.textbox1


        y=additem1.textbox2


        obj = recordoverride(x,y)

        self.main_layout.clear_widgets()
        f = additem1()
        f.run()

    def on_button_press2(self, instance):
        self.main_layout.clear_widgets()
        f = additem1()
        f.run()
class itemslist():
    def __init__(self):

        self.output = list()
        wb = openpyxl.load_workbook("datafile.xlsx")
        sheet = wb['Sheet1']  # wb.get_sheet_names())
        columnvalue=1
        for i in range(1, sheet.max_row + 1):

            if sheet.cell(row=i, column=1).value == None:
                pass
            else:
                self.output.append(sheet.cell(row=i, column=1).value)
        try:
            self.output.remove(None)
        except:
            pass
        print(self.output)

    def get(self):
        return self.output


class record_add():
    def __init__(self, given_str1, given_str2):
        ctr = 0
        given_str1 = str(given_str1)
        given_str2 = str(given_str2)




        wb = openpyxl.load_workbook("datafile.xlsx")
        sheet = wb['Sheet1']  # wb.get_sheet_names())


        columnvalue = 1
        for i in range(1, sheet.max_row + 1):
            if ((sheet.cell(row=i, column=columnvalue).value == given_str1)):
                ctr = 1

                cell = sheet.cell(row=i, column=1)
                cell.value = given_str1

                cell = sheet.cell(row=i, column=2)
                cell.value = given_str2


                break
            else:
                pass
        if (ctr == 0):

            for i in range(1, sheet.max_row + 2):
                # print('maxrow',sheet.max_row)

                if str(sheet.cell(row=i, column=1).value) == 'None':
                    cell = sheet.cell(row=i, column=1)
                    cell.value = given_str1

                    cell = sheet.cell(row=i, column=2)
                    cell.value = given_str2


                    break


        else:
            pass
        wb.save("datafile.xlsx")


    def get(self):
        return self.output
class record_availability_checker():
    def __init__(self,given_str):
        ctr=0
        given_str=str(given_str)
        given_str=given_str.strip()

        wb = openpyxl.load_workbook("datafile.xlsx")
        sheet = wb['Sheet1']  # wb.get_sheet_names())

        columnvalue=1
        for i in range(1, sheet.max_row + 1):
            if ((sheet.cell(row=i, column=columnvalue).value == given_str)):
                ctr = 1

                self.output = sheet.cell(row=i, column=2).value

                break
            else:
                pass

        if (ctr == 0):

            self.output = ""

    def get(self):
        return self.output
class recordoverride():
    def __init__(self,given_str1,given_str2):

        given_str1=str(given_str1)
        given_str2 = str(given_str2)
        wb = openpyxl.load_workbook("datafile.xlsx")
        sheet = wb['Sheet1']  # wb.get_sheet_names())
        print('recordoverride str 1 = ',given_str1)
        for i in range(1, sheet.max_row + 1):
            if ((sheet.cell(row=i, column=1).value == given_str1)):

                cell = sheet.cell(row=i, column=1)
                cell.value = given_str1
                cell = sheet.cell(row=i, column=2)
                cell.value = given_str2


                break
            else:
                pass

        wb.save("datafile.xlsx")

class words_filter():
    def __init__(self,str1):
                    self.lis3 = ""
                    for i in str1:
                        if i !='\n' and i !='\t':
                            self.lis3=self.lis3+i
                        else:
                            pass

    def output(self):


        return self.lis3
if __name__ == "__main__":
    app = MainApp()
    app.run()
